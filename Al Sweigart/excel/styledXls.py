import openpyxl
from openpyxl.styles import Font, Alignment


font = Font(name='Times New Roman',
            size=24,
            bold=True,
            italic=True,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
alignment = Alignment(horizontal='left',
                      vertical='bottom',
                      shrinkToFit=False,
                      wrapText=False,
                      indent=0)

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Здравствуй, мир!'
ws['A5'] = 'Пока!'
max_row = ws.max_row
for row in range(1, max_row + 1):
    cel = 'A' + str(row)
    ws[cel].font = font
    ws[cel].alignment = alignment
ws.column_dimensions['A'].width = 40
wb.save('styled.xlsx')
