# Программа корректирует цены в электронной таблице produceSales.xlsx.


import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

# Типы продукции и их обновленные цены.
price_updates = {'Lemon': 3.07,
                 'Celery': 1.19,
                 'Garlic': 1.27}
# Создать цикл по строкам и обновить цены.
maxRow = sheet.max_row
for rowNum in range(2, maxRow):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in price_updates:
        sheet.cell(row=rowNum, column=2).value = price_updates[produceName]
wb.save('updateProduceSales.xlsx')
