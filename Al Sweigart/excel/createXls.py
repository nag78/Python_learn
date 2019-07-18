import openpyxl


wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)

# Переименование страниц
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.create_sheet(title='Первый лист', index=0)


# Запись значения в ячейки
sheet = wb[wb.sheetnames[0]]
sheet['A1'] = 'Здравствуй, мир!'
print(sheet['A1'].value)


wb.save('example_copy.xlsx')
