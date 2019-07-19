import openpyxl


wb = openpyxl.Workbook()
ws = wb.active
ws.merge_cells('A1:D3')
ws['A1'] = 'Объеденены двенадцать ячеек.'
ws.merge_cells('C5:E5')
ws['C5'] = 'Объеденены три ячейки.'

wb.save('merged.xlsx')
