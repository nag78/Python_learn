import openpyxl


wb = openpyxl.load_workbook('merged.xlsx')
ws = wb.active
ws.unmerge_cells('A1:D3')
ws.unmerge_cells('C5:E5')
wb.save('merged.xlsx')
