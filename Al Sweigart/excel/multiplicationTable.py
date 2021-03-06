import sys
import openpyxl

# откроем Рабочую книгу
wb = openpyxl.Workbook()
ws = wb.active
# получение factor через аргумент строки
argument = ''.join(sys.argv[1:])
if argument == '':
    print('Нет аргументов строки.')
else:
    factor = int(argument)

    # Разметка таблицы

    for i in range(1, factor + 1):
        _ = ws.cell(row=i+1, column=1, value=i)
        _ = ws.cell(row=1, column=i+1, value=i)


    # Генератор таблицы
    tabl = [[i*j for j in range(1, factor + 1)] for i in range(1, factor + 1)]

    col = 2
    row = 2
    for subtabl in tabl:
        for index, value in enumerate(subtabl):
            ws.cell(column=col+index, row=row).value = value
        row += 1

    filename = 'multiple' + str(factor) + '.xlsx'
    wb.save(filename)
