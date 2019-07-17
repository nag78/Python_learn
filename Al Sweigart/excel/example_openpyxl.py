# -*- coding: cp1251 -*-

from openpyxl import load_workbook
wb = load_workbook(filename='example.xlsx')
sheet = wb.active
c = sheet['B1']
print('Строка ' + str(c.row) + ' Столбец ' + str(c.column) +
      ' : ' + str(c.value))
time = sheet['A1'].value
print(time)

print('Ячейка ' + str(c.coordinate) + ' : ' + str(c.value))

print(str(sheet.cell(row=1, column=2)))

for i in range(1, 8, 3):
    print(i, sheet.cell(row=i, column=2).value)
hr = sheet.calculate_dimension()
print(str(hr))

hr2 = sheet.max_row
hc = sheet.max_column
print(str(hr2) + ' : ' + str(hc))

ts = tuple(sheet['A1':'C7'])
# print(str(ts))
for rowOfCellObjects in sheet['A1':'C7']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---Конец строки---')
