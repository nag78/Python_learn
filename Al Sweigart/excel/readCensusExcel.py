# -*- coding: cp1251 -*-

import openpyxl
import Lib.pprint


print('Открытие рабочей книги...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb[wb.sheetnames[0]]
countyData = {}
print('Чтение строк...')
for row in range(2, sheet.max_row + 1):
    # В каждой строке содержатся данные
    # для одного переписного района.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    # Гарантия существования ключа для
    # данного штата.
    countyData.setdefault(state, {})
    # Гарантия существования ключа для
    # данного округа данного штата.
    countyData[state].setdefault(county, {'tracts': 0,
                                 'pop': 0})
    # Каждая строка представляет собой
    # один переписной район, поэтому
    # инкреминировать единицу.
    countyData[state][county]['tracts'] += 1
    # Увеличить численность населения
    # округа на численность населения
    # переписного района.
    countyData[state][county]['pop'] += int(pop)

# Открыть новый текстовый файл и
# записать в него содержимое
# словоря countyData.
print('Запись результатов...')
resultFile = open('county2010.py', 'w')
resultFile.write('allData = ' + Lib.pprint.pformat(countyData))
resultFile.close()
print('Готово!')
