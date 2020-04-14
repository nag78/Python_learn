import sys
from openpyxl import load_workbook
import openpyxl

argument = ''.join(sys.argv[1])
filein = str(argument)
print(filein)
wb = load_workbook(filename=filein)
ws = wb.active
wb1 = openpyxl.Workbook()
ws1 = wb1.active

rows = 1
cols = 1
for row in ws.rows:
    rows += 1

for col in ws.columns:
    cols += 1
tabl = [[ws.cell(column=i, row=j).value
        for j in range(1, rows)] for i in range(1, cols)]
# print(tabl)

tabl1 = list(zip(*tabl))
# print(tabl1)
col = 1
row = 1
for subtable in tabl1:
    for index, value in enumerate(subtable):
        ws1.cell(column=col, row=row+index).value = value
    col += 1


fileout = 'inv' + filein
print(fileout)
wb1.save(filename=fileout)
