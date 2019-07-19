import openpyxl


wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Высокая строка'
ws['B2'] = 'Широкий столбец'
ws.row_dimensions[1].height = 70
ws.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')
