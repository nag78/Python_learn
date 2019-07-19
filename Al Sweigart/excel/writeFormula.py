import openpyxl


wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 200
ws['A2'] = 300
ws['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')
